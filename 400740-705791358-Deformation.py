import requests
import json
import openpyxl as op
import numpy as np
import os

from xlrd import open_workbook
from xlutils.copy import copy


# ------------------------------窗口输出信息------------------------------
def configprint(vae, did, point):
    # 计数循环，窗口递增输出数据流信息
    count = 1
    print('设备ID' + '\t\t\t\t\t' + '变量名称' + '\t\t\t\t' + '最新数据' + '\t\t\t\t' + '时间')
    for index, values in enumerate(point):
        count += 1
        time = str(values.get('at', ''))
        temperature = str(values.get('value', ''))
        print(did + '\t\t\t\t' + vae + '\t\t\t\t' + temperature + '\t\t\t\t' + time)


# ------------------------------获取onenet设备数据------------------------------
def setconfig(did, api, vae, startid, streamHttp, pointHttp):
    # pid:项目名称,vae:变量名称,did:设备ID,point:数据信息,tit:设备名称,eid:excel名称及路径,sid:sheet名(形参下同)
    payload = {'datastream_id': vae, 'start': startid, 'limit': 1}
    # 参数设置
    headers = {'api-key': api}
    # # # 设备历史数据API
    url_point = pointHttp + did + '/datapoints?'
    Point = requests.get(url_point, headers=headers, params=payload)

    # 从设备历史数据中取出数据流中数据信息
    temp = str(Point.text)
    Jtemp = json.loads(temp)
    data = Jtemp['data']
    datastreams = data['datastreams']
    for index, values in enumerate(datastreams):
        point = values.get('datapoints', '')

    return [point]



# ------------------------------数据导入excel------------------------------
def append_Excel(point, eid):
    # writeExcel(point, excelId0, sheetId0, count0)
    # 打开excel
    r_xls = open_workbook(eid)  # 读取excel文件
    row = r_xls.sheets()[0].nrows - 1  # 获取已有的行数
    print(row)
    excel = copy(r_xls)  # 将xlrd的对象转化为xlwt的对象
    table = excel.get_sheet(0)  # 获取要操作的sheet

    for index, values in enumerate(point):
        row += 1
        time = str(values.get('at', ''))  # time代表数据信息对应时间'at'，temp代表数据信息'value'
        temp = str(values.get('value', ''))  # 随循环递增exlce表内容，row代表行，column代表列，value代表要添加的信息

        if time != "":
            table.write(row, 0, time)  # 括号内分别为行数、列数、内容
            table.write(row, 1, temp)

    excel.save(eid)  # 保存并覆盖文件
    print('追加成功！')


def writeExcel(eid, sid):
    # writeExcel(point, excelId0, sheetId0, count0)
    # 创建excel表
    ws = op.Workbook()
    # 创建sheet表单
    wb = ws.create_sheet(sid)
    # 表头信息
    wb.cell(row=1, column=1, value='时间')
    wb.cell(row=1, column=2, value=variAble)
    ws.save(eid)


# ------------------------------main()------------------------------
if __name__ == "__main__":
    project_name = '膨胀量测试'
    productId = '400740'
    device_name = 'LVDT2'
    deviceId = '705791358'  # 设备ID
    APIKey = '0WSUNK4bOl3mqNIw9Lab6D6GDIg='  # API配置参数
    variAble = 'Deformation'  # 变量名称
    streamHttp = 'http://api.heclouds.com/devices/'  # 对应url使用参数
    pointHttp = 'http://api.heclouds.com/devices/'  # 对应url使用参数

    M0 = np.empty([2030, 12], dtype=int)
    M0[2021, :] = [31, 29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]  # 开始年份每个月具有的天数
    M0[2022, :] = [30, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]  # 结束年份每个月具有的天数
    M0[2023, :] = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]

    start_year = 2022
    start_month = 7  # 1-12
    start_day = 12  # 大于1小于等于当月天数
    start_hour = 0  # 0-23小时制
    start_mine = 0  # 0-59
    start_second = 0  # 0-59

    Interval_hour = 1  # 1-24
    Interval_mine = 1  # 1-60
    Interval_second = 60  # 1-60
    #######################################################################################################################
    filename = 'D:\数据\产品(' + project_name + ')' + productId + '\设备(' + device_name + ')' + deviceId + '\变量名' + variAble
    sheetId = 'Sheet'  # sheet名称
    # 文件路径
    word_name = os.path.exists(filename)
    # 判断文件是否存在：不存在创建
    if not word_name:
        os.makedirs(filename)
    excelad = filename + '\数据'


    for ij in range(start_year, start_year+2):
        if ij == start_year:  # 年
            current_month = start_month
        else:
            current_month = 1

        year = str(ij) + '-'

        M = M0[ij, :] #取二维数组的行

        for i in range(current_month, 13):  # 控制月
            # 2021-09-24T16:19:44
            if ij == start_year and i == start_month:  # 年
                current_day = start_day
            else:
                current_day = 1

            if i < 10:  # 控制月
                month = '0' + str(i) + '-'
            else:
                month = str(i) + '-'

            for j in range(current_day, M[i - 1]+1):  # 控制天
                # 调用configSet设置参数

                if ij == start_year and i == start_month and j == start_day:  # 年
                    current_hour = start_hour
                else:
                    current_hour = 0

                if j < 10:
                    day = '0' + str(j)
                else:
                    day = str(j)

                count = 1  # 控制文件

                for k in range(start_hour, 24, Interval_hour):  # 控制时
                    # 调用configSet设置参数
                    if ij == start_year and i == start_month and j == start_day and k == start_hour:  # 年
                        current_mine = start_mine
                    else:
                        current_mine = 0

                    if k < 10:
                        hour = 'T0' + str(k)
                    else:
                        hour = 'T' + str(k)

                    for m in range(start_mine, 60, Interval_mine):  # 控制分钟#for i in range(5,10,2):   间隔为2
                        if ij == start_year and i == start_month and j == start_day and k == start_hour and m == start_mine:  # 年
                            current_second = start_second
                        else:
                            current_second = 0

                        if m < 10:
                            mine = ':0' + str(m)
                        else:
                            mine = ':' + str(m)

                        for n in range(start_second, 60, Interval_second):  # 控制分钟#for i in range(5,10,2):   间隔为2
                            if n < 10:
                                second = ':0' + str(n)
                            else:
                                second = ':' + str(n)

                            excelId0 = excelad + productId + '-' + deviceId + '-' + variAble + '-' + year + month + day + '.xls'
                            startId0 = year + month + day + hour + mine + second
                            print(startId0)

                            if count == 1:
                                writeExcel(excelId0, sheetId)
                                count = count + 1

                            configSet = setconfig(deviceId, APIKey, variAble, startId0, streamHttp, pointHttp)
                            point = configSet[0]  # 数据流中数据信息
                            #configprint(variAble, deviceId, point)
                            append_Excel(point, excelId0)
