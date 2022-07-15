import requests
import json
import openpyxl as op
import pandas as pd
from matplotlib import pyplot as plt
import configparser
import os

iii = 1
print(iii)

# ------------------------------窗口输出信息------------------------------
def configprint(vae, did, point):
    # 计数循环，窗口递增输出数据流信息
    count = 1
    print('设备ID' + '\t\t\t\t\t' + '变量名称' + '\t\t\t\t\t\t' + '最新数据' + '\t\t\t\t' + '时间')
    for index, values in enumerate(point):
        count += 1
        time = str(values.get('at', ''))
        temperature = str(values.get('value', ''))
        print(did + '\t\t\t\t' + vae + '\t\t\t\t' + temperature + '\t\t\t\t' + time)
# ------------------------------获取onenet设备数据------------------------------
def setconfig(did, api, vae, startid, streamHttp, pointHttp):
    # pid:项目名称,vae:变量名称,did:设备ID,point:数据信息,tit:设备名称,eid:excel名称及路径,sid:sheet名(形参下同)
    # url配置参数
    # payload = {'start': startid, 'limit': 6000}
    # http: // api.heclouds.com / devices / 8029377 / datapoints?datastream_id = ds & start = 2017 - 01 - 01
    # T00: 00:00 & limit = 100
    # HTTP / 1.1

    payload = {'datastream_id': vae, 'start': startid, 'limit': 40}  # 参数设置
    headers = {'api-key': api}
    # streamHttp = http: // api.heclouds.com / devices /
    # 设备详情API
    url_stream = streamHttp + did  # +'/datapoints?datastream_id ='+vae
    # 从设备详情信息中取出设备号数据(title)
    Title = requests.get(url_stream, headers=headers)
    temp = str(Title.text)
    Jtemp = json.loads(temp)
    data = Jtemp['data']
    keys = data['keys']
    for index, values in enumerate(keys):
        title = values.get('title', '')
    # 设备历史数据API
    url_point = pointHttp + did + "/datapoints?"
    Point = requests.get(url_point, headers=headers, params=payload)
    # 从设备历史数据中取出数据流中数据信息
    temp = str(Point.text)
    Jtemp = json.loads(temp)
    data = Jtemp['data']
    datastreams = data['datastreams']
    for index, values in enumerate(datastreams):
        point = values.get('datapoints', '')
    return [title, point]


# ------------------------------数据导入excel------------------------------
def writeExcel(point, eid, sid):
    # 创建excel表
    ws = op.Workbook()
    # 创建sheet表单
    wb = ws.create_sheet(sid)
    # 表头信息
    # wb.cell(row=1, column=1, value='项目名称')
    # wb.cell(row=1, column=2, value='设备名称')
    wb.cell(row=1, column=1, value='时间')
    wb.cell(row=1, column=2, value='最新数据')

    # 计数器，代表行数
    count = 1
    # 循环数据信息，每次循环一个字典，计数+1
    for index, values in enumerate(point):
        count += 1
        # time代表数据信息对应时间'at'，temp代表数据信息'value'
        time = str(values.get('at', ''))
        temp = str(values.get('value', ''))
        # 随循环递增exlce表内容，row代表行，column代表列，value代表要添加的信息
        # wb.cell(row=count, column=1, value='项目' + pid)
        # wb.cell(row=count, column=2, value='设备' + tit)
        wb.cell(row=count, column=1, value=time)
        wb.cell(row=count, column=2, value=temp)
        # 保存表格
        ws.save(eid)
    ws.close()





# ------------------------------main()------------------------------
if __name__ == "__main__":
    projectId = 'python实战'  # 项目名称
    variAble = 'Deformation'  # 变量名称
    deviceId = '705791358'  # 设备ID
    APIKey = '0WSUNK4bOl3mqNIw9Lab6D6GDIg='  # API配置参数
    sheetId = 'Sheet1'  # sheet名称
    streamHttp = 'http://api.heclouds.com/devices/'  # 对应url使用参数
    pointHttp = 'http://api.heclouds.com/devices/'  # 对应url使用参数
    excelad = 'D:\大论文\物联网数据/'
    year1 = '2021-'
    year2 = '2022-'
    #######################################################################################################################
    M1 = [31, 29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    M2 = [30, 28, 31, 30, 31, 30, 31]#, 31, 30, 31, 30, 31]
    N_year=2
    N = 1
    for ij in range(N_year-1):
       ij=1
       if ij == 0:  # 控制月
           year = year1
           M = M1
           current = 7
       else:
           year = year2
           M = M2
           current = 1

       for i in range(current-1, 12):#控制月
            #2021-09-24T16:19:44
            if i < 9:#控制月
                month = '0' + str(i + 1) + '-'
            else:
                month = str(i + 1) + '-'
            for j in range(M[i]):#控制天
                # 调用configSet设置参数
                if j < 9:
                    day = '0' + str(j + 1)
                else:
                    day = str(j + 1)
                for k in range(24):#控制时
                    # 调用configSet设置参数
                    if k < 10:
                        hour = 'T0' + str(k)
                    else:
                        hour = 'T' + str(k)

                    sheetId0 = 'Sheet1'
                    N = N+1
                    #excelId0 = excelad + deviceId + variAble + year + str(N) + '.xlsx'
                    excelId0 = excelad + deviceId + variAble + year + month + day + hour + '.xlsx'
                    startId0 = year + month + day + hour + ':00:00'
                    #tartId0 = year + month + day + 'T00:00:00'
                    #print(excelId0)
                    print(startId0)

                    configSet = setconfig(deviceId, APIKey, variAble, startId0, streamHttp, pointHttp)
                    title = configSet[0]  # 设备名称
                    point = configSet[1]  # 数据流中数据信息
                    for index, values in enumerate(point):
                        time = str(values.get('at', ''))
                        temperature = float((values.get('value', '')))
                        if (time != ""):
                            # 数据对应时间不为空，即可执行窗口输出，写入excel
                            #configprint(variAble, deviceId, point)
                            writeExcel(point, excelId0, sheetId0)
                            break
                        else:
                            print("error:NO POINT!")
                        break



